import json
import os
import re
from pathlib import Path

import pandas as pd

Path("tmp/matplotlib").mkdir(parents=True, exist_ok=True)
os.environ.setdefault("MPLCONFIGDIR", str(Path("tmp/matplotlib")))

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt

from src.commands.generate_insights import _build_report
from src.commands.summarize import _build_summary
from src.core.session_memory import load_session_memory
from src.core.workbook_manager import WorkbookError, read_sheet, resolve_sheet_name


DASHBOARD_OUTPUT_DIR = Path("outputs") / "dashboards"
MONTH_ORDER = {
    "january": 1,
    "jan": 1,
    "february": 2,
    "feb": 2,
    "march": 3,
    "mar": 3,
    "april": 4,
    "apr": 4,
    "may": 5,
    "june": 6,
    "jun": 6,
    "july": 7,
    "jul": 7,
    "august": 8,
    "aug": 8,
    "september": 9,
    "sep": 9,
    "sept": 9,
    "october": 10,
    "oct": 10,
    "november": 11,
    "nov": 11,
    "december": 12,
    "dec": 12,
}
METRIC_CANDIDATES = (
    "Revenue",
    "Sales",
    "Amount",
    "Total",
    "Profit",
    "Income",
    "Value",
)


def _current_session_file():
    current_file = load_session_memory().get("current_file")
    if isinstance(current_file, str) and current_file.lower().endswith(".xlsx"):
        return current_file
    return None


def _current_session_sheet():
    current_sheet = load_session_memory().get("current_sheet")
    if isinstance(current_sheet, str) and current_sheet:
        return current_sheet
    return None


def _dashboard_dir(file_path, sheet_name=None):
    source = Path(file_path)
    DASHBOARD_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if sheet_name:
        return DASHBOARD_OUTPUT_DIR / f"{source.stem}_{_slugify(sheet_name)}_dashboard"
    return DASHBOARD_OUTPUT_DIR / f"{source.stem}_dashboard"


def _normalize_name(value):
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def _slugify(value):
    text = str(value or "dashboard").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_") or "dashboard"


def _match_column(df, requested_column):
    if not requested_column:
        return None

    requested = _normalize_name(requested_column)
    for column in df.columns:
        if _normalize_name(column) == requested:
            return column

    return None


def _find_column(df, candidates):
    for candidate in candidates:
        column = _match_column(df, candidate)
        if column is not None:
            return column
    return None


def _numeric_columns(df):
    return list(df.select_dtypes(include="number").columns)


def _categorical_columns(df):
    numeric = set(_numeric_columns(df))
    return [
        column
        for column in df.columns
        if column not in numeric and not pd.api.types.is_datetime64_any_dtype(df[column])
    ]


def _choose_metric_column(df, target_column=None):
    numeric_columns = _numeric_columns(df)
    requested_column = _match_column(df, target_column)
    if requested_column in numeric_columns:
        return requested_column

    for candidate in METRIC_CANDIDATES:
        column = _match_column(df, candidate)
        if column in numeric_columns:
            return column

    return numeric_columns[0] if numeric_columns else None


def _choose_group_column(df, group_by=None):
    requested_column = _match_column(df, group_by)
    if requested_column is not None:
        return requested_column

    category_column = _find_column(df, ("Category", "Product", "Segment", "Region"))
    if category_column is not None:
        return category_column

    categorical_columns = _categorical_columns(df)
    return categorical_columns[0] if categorical_columns else None


def _choose_month_column(df):
    month_column = _find_column(df, ("Month", "Date", "Order Date", "Transaction Date"))
    if month_column is not None:
        return month_column

    for column in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[column]):
            return column

    return None


def _month_sort_key(value):
    text = str(value).strip().lower()
    if text in MONTH_ORDER:
        return (0, MONTH_ORDER[text])

    parsed = pd.to_datetime(value, errors="coerce")
    if not pd.isna(parsed):
        return (1, parsed)

    return (2, text)


def _chart_path(charts_dir, title, chart_type):
    charts_dir.mkdir(parents=True, exist_ok=True)
    return charts_dir / f"{_slugify(title)}_{chart_type}_chart.png"


def _chart_record(title, chart_type, path, x_column=None, y_column=None):
    return {
        "title": title,
        "chart_type": chart_type,
        "file": path.as_posix(),
        "x_column": str(x_column) if x_column is not None else None,
        "y_column": str(y_column) if y_column is not None else None,
    }


def _plot_series(series, chart_type, title, x_label, y_label, output_path):
    if series is None or series.empty:
        return False

    fig, ax = plt.subplots(figsize=(9, 5))
    if chart_type == "line":
        series.plot(kind="line", ax=ax, marker="o")
    elif chart_type == "pie":
        series.plot(kind="pie", ax=ax, autopct="%1.1f%%")
        ax.set_ylabel("")
    else:
        series.plot(kind="bar", ax=ax)

    ax.set_title(title)
    if chart_type != "pie":
        ax.set_xlabel(str(x_label))
        ax.set_ylabel(str(y_label))
    fig.tight_layout()
    fig.savefig(output_path)
    plt.close(fig)
    return True


def _plot_histogram(df, metric_column, charts_dir):
    if metric_column is None:
        return None

    series = pd.to_numeric(df[metric_column], errors="coerce").dropna()
    if series.empty:
        return None

    title = f"{metric_column} Distribution"
    output_path = _chart_path(charts_dir, title, "histogram")
    fig, ax = plt.subplots(figsize=(8, 5))
    series.plot(kind="hist", ax=ax, bins=min(10, max(3, len(series))))
    ax.set_title(title)
    ax.set_xlabel(str(metric_column))
    ax.set_ylabel("Frequency")
    fig.tight_layout()
    fig.savefig(output_path)
    plt.close(fig)
    return _chart_record(title, "histogram", output_path, x_column=metric_column)


def _group_metric_series(df, group_column, metric_column):
    if group_column is None or metric_column is None:
        return None

    chart_df = df[[group_column, metric_column]].copy()
    chart_df[metric_column] = pd.to_numeric(chart_df[metric_column], errors="coerce")
    chart_df = chart_df.dropna(subset=[group_column, metric_column])
    if chart_df.empty:
        return None

    return (
        chart_df.groupby(group_column, dropna=False)[metric_column]
        .sum()
        .sort_values(ascending=False)
        .head(10)
    )


def _month_metric_series(df, month_column, metric_column):
    if month_column is None or metric_column is None:
        return None

    chart_df = df[[month_column, metric_column]].copy()
    chart_df[metric_column] = pd.to_numeric(chart_df[metric_column], errors="coerce")
    chart_df = chart_df.dropna(subset=[month_column, metric_column])
    if chart_df.empty:
        return None

    grouped = chart_df.groupby(month_column, sort=False, dropna=False)[metric_column].sum()
    sorted_items = sorted(grouped.items(), key=lambda item: _month_sort_key(item[0]))
    return pd.Series(
        [value for _, value in sorted_items],
        index=[str(label) for label, _ in sorted_items],
    )


def _create_charts(df, dashboard_dir, target_column=None, group_by=None):
    charts_dir = dashboard_dir / "charts"
    charts = []
    metric_column = _choose_metric_column(df, target_column)
    group_column = _choose_group_column(df, group_by)
    month_column = _choose_month_column(df)

    grouped = _group_metric_series(df, group_column, metric_column)
    if grouped is not None and not grouped.empty:
        title = f"{metric_column} by {group_column}"
        output_path = _chart_path(charts_dir, title, "bar")
        if _plot_series(grouped, "bar", title, group_column, metric_column, output_path):
            charts.append(
                _chart_record(
                    title,
                    "bar",
                    output_path,
                    x_column=group_column,
                    y_column=metric_column,
                )
            )

        pie_title = f"{metric_column} Share by {group_column}"
        pie_path = _chart_path(charts_dir, pie_title, "pie")
        if _plot_series(grouped.head(6), "pie", pie_title, group_column, metric_column, pie_path):
            charts.append(
                _chart_record(
                    pie_title,
                    "pie",
                    pie_path,
                    x_column=group_column,
                    y_column=metric_column,
                )
            )

    monthly = _month_metric_series(df, month_column, metric_column)
    if monthly is not None and len(monthly) >= 2:
        title = f"{metric_column} Trend by {month_column}"
        output_path = _chart_path(charts_dir, title, "line")
        if _plot_series(monthly, "line", title, month_column, metric_column, output_path):
            charts.append(
                _chart_record(
                    title,
                    "line",
                    output_path,
                    x_column=month_column,
                    y_column=metric_column,
                )
            )

    histogram = _plot_histogram(df, metric_column, charts_dir)
    if histogram:
        charts.append(histogram)

    return charts


def _write_text_file(path, content):
    with path.open("w", encoding="utf-8") as file:
        file.write(content)


def _write_manifest(path, manifest):
    with path.open("w", encoding="utf-8") as file:
        json.dump(manifest, file, indent=2)


def _write_excel_report(report_path, df, summary, insights, findings, recommendations, charts):
    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        pd.DataFrame({"Summary": summary.splitlines()}).to_excel(
            writer,
            sheet_name="Summary",
            index=False,
        )
        pd.DataFrame({"Insight Report": insights.splitlines()}).to_excel(
            writer,
            sheet_name="Insights",
            index=False,
        )
        pd.DataFrame(charts or [{"title": "No charts created"}]).to_excel(
            writer,
            sheet_name="Charts",
            index=False,
        )
        df.head(50).to_excel(writer, sheet_name="Data Preview", index=False)

    try:
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image as ExcelImage
        from openpyxl.styles import Font

        workbook = load_workbook(report_path)
        dashboard_sheet = workbook.create_sheet("Dashboard", 0)
        dashboard_sheet["A1"] = "Auto Dashboard"
        dashboard_sheet["A1"].font = Font(bold=True, size=16)
        dashboard_sheet["A3"] = "Dataset"
        dashboard_sheet["B3"] = f"{len(df)} rows x {len(df.columns)} columns"
        dashboard_sheet["A4"] = "Charts created"
        dashboard_sheet["B4"] = len(charts)

        dashboard_sheet["A6"] = "Key Findings"
        dashboard_sheet["A6"].font = Font(bold=True)
        row = 7
        for finding in findings[:6]:
            dashboard_sheet.cell(row=row, column=1, value=f"- {finding}")
            row += 1

        row += 1
        dashboard_sheet.cell(row=row, column=1, value="Recommendations")
        dashboard_sheet.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        for recommendation in recommendations[:6]:
            dashboard_sheet.cell(row=row, column=1, value=f"- {recommendation}")
            row += 1

        charts_sheet = workbook["Charts"]
        image_row = max(len(charts) + 4, 6)
        for chart in charts:
            try:
                image = ExcelImage(chart["file"])
                image.width = 480
                image.height = 270
                charts_sheet.add_image(image, f"A{image_row}")
                image_row += 18
            except Exception:
                continue

        workbook.save(report_path)
    except Exception:
        pass


def _build_manifest(file_path, sheet_name, dashboard_dir, summary_file, insights_file, report_file, charts):
    return {
        "source_file": str(file_path),
        "sheet_name": sheet_name,
        "dashboard_dir": dashboard_dir.as_posix(),
        "summary_file": summary_file.as_posix(),
        "insights_file": insights_file.as_posix(),
        "dashboard_file": report_file.as_posix(),
        "charts": charts,
    }


def build_dashboard(file_path, target_column=None, group_by=None, sheet_name=None):
    """Build a dashboard folder with summary, insights, charts, and an Excel report."""
    file_path = file_path or _current_session_file()
    if not file_path:
        message = "No input file provided and no current session file found."
        print(message)
        return {
            "status": "error",
            "output_file": None,
            "message": message,
        }

    source = Path(file_path)
    if not source.exists():
        message = f"File not found: {file_path}"
        print(message)
        return {
            "status": "error",
            "output_file": None,
            "message": message,
        }

    try:
        resolved_sheet = resolve_sheet_name(
            file_path,
            requested_sheet=sheet_name,
            session_sheet=_current_session_sheet(),
        )
        df = read_sheet(source, resolved_sheet)
    except WorkbookError as error:
        message = str(error)
        print(message)
        return {
            "status": "error",
            "output_file": None,
            "message": message,
        }
    except ValueError as error:
        message = f"Invalid file: {error}"
        print(message)
        return {
            "status": "error",
            "output_file": None,
            "message": message,
        }
    except Exception as error:
        message = f"Error reading file: {error}"
        print(message)
        return {
            "status": "error",
            "output_file": None,
            "message": message,
        }

    dashboard_dir = _dashboard_dir(file_path, resolved_sheet if sheet_name else None)
    dashboard_dir.mkdir(parents=True, exist_ok=True)
    summary_file = dashboard_dir / "summary.txt"
    insights_file = dashboard_dir / "insights.txt"
    report_file = dashboard_dir / "dashboard.xlsx"
    manifest_file = dashboard_dir / "manifest.json"

    summary = _build_summary(df)
    insights, findings, recommendations = _build_report(
        df,
        file_path,
        target_column=target_column,
        group_by=group_by,
    )
    charts = _create_charts(
        df,
        dashboard_dir,
        target_column=target_column,
        group_by=group_by,
    )

    _write_text_file(summary_file, summary)
    _write_text_file(insights_file, insights)
    _write_excel_report(report_file, df, summary, insights, findings, recommendations, charts)
    manifest = _build_manifest(
        file_path,
        resolved_sheet,
        dashboard_dir,
        summary_file,
        insights_file,
        report_file,
        charts,
    )
    _write_manifest(manifest_file, manifest)

    message = f"Dashboard built successfully: {dashboard_dir.as_posix()}"
    print(message)
    print(f"Dashboard report: {report_file.as_posix()}")
    print(f"Sheet used: {resolved_sheet}")
    print("")
    print("Dashboard Summary:")
    print(f"- Rows: {len(df)}")
    print(f"- Columns: {len(df.columns)}")
    print(f"- Charts created: {len(charts)}")
    if findings:
        print(f"- Top finding: {findings[0]}")

    return {
        "status": "success",
        "output_file": dashboard_dir.as_posix(),
        "message": message,
        "result_summary": "Dashboard built successfully.",
        "dashboard_file": report_file.as_posix(),
        "summary_file": summary_file.as_posix(),
        "insights_file": insights_file.as_posix(),
        "manifest_file": manifest_file.as_posix(),
        "charts": charts,
        "sheet_name": resolved_sheet,
    }
